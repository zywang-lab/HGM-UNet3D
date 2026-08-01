import os
import torch
import numpy as np
from tqdm import tqdm
import torch.optim as optim
from openpyxl import Workbook
from models.HGM_Unet3D import HGM_UNet3D
from monai.losses import DiceLoss
from torch.utils.data import DataLoader
from generators.image_label_generator import Image_Label_train, Image_Label_valid



def eval(net, dataloader):

    net.eval()
    loss_valid = []
    dice_scores = []
    iou_scores = []
    voe_scores = []
    precision_scores = []
    recall_scores = []
    accuracy_scores = []

    with torch.no_grad():
        for (image, label) in dataloader:
            image = image.to(device)
            label = label.to(device)
            label_pred = net(image)
            loss = dsc_loss(label_pred, label)
            loss_valid.append(loss.item())
            pred_prob = torch.sigmoid(label_pred)
            dice_score = calculate_dice(pred_prob, label)
            dice_scores.append(dice_score)
            iou_score = calculate_iou(pred_prob, label)
            iou_scores.append(iou_score)
            voe_score = calculate_voe(pred_prob, label)
            voe_scores.append(voe_score)
            precision_score = calculate_precision(pred_prob, label)
            precision_scores.append(precision_score)
            recall_score = calculate_recall(pred_prob, label)
            recall_scores.append(recall_score)
            accuracy_score = calculate_accuracy(pred_prob, label)
            accuracy_scores.append(accuracy_score)
    return np.mean(loss_valid), np.mean(dice_scores), np.mean(iou_scores), np.mean(voe_scores), np.mean(
        precision_scores), np.mean(recall_scores), np.mean(accuracy_scores)


def calculate_dice(pred, target):
    smooth = 1e-5
    pred = (pred > 0.5).float()
    intersection = (pred * target).sum()
    union = pred.sum() + target.sum()
    dice = (2. * intersection + smooth) / (union + smooth)
    return dice.item()

def calculate_voe(pred, target):
    smooth = 1e-5
    pred = (pred > 0.5).float()
    intersection = (pred * target).sum()
    union = pred.sum() + target.sum() - intersection
    iou = (intersection + smooth) / (union + smooth)
    return (1 - iou).item()


def calculate_iou(pred, target):
    smooth = 1e-5
    pred = (pred > 0.5).float()
    intersection = (pred * target).sum()
    union = pred.sum() + target.sum() - intersection
    iou = (intersection + smooth) / (union + smooth)
    return iou.item()


def calculate_precision(pred, target):
    smooth = 1e-5
    pred = (pred > 0.5).float()
    true_positive = (pred * target).sum()
    predicted_positive = pred.sum()
    precision = (true_positive + smooth) / (predicted_positive + smooth)
    return precision.item()


def calculate_recall(pred, target):
    smooth = 1e-5
    pred = (pred > 0.5).float()
    true_positive = (pred * target).sum()
    actual_positive = target.sum()
    recall = (true_positive + smooth) / (actual_positive + smooth)
    return recall.item()

def calculate_accuracy(pred, target):
    smooth = 1e-5
    pred = (pred > 0.5).float()
    correct = (pred == target).float().sum()
    total = target.numel()
    accuracy = (correct + smooth) / (total + smooth)
    return accuracy.item()


#################################################################

epochs = 800
real_batch_size = 8
accumulation_steps = 4
effective_batch_size = real_batch_size * accumulation_steps

init_learning_rate = 0.0005
learning_rate_patience = 12
learning_rate_factor = 0.5
modelsave_path = "..\\modelsave\\EViT_ResUNet3D_Gated_Bidirection_layer3_depth2\\"
image_patch_folder = "E:\\Datasets\\EM710\\images_patches\\"
label_patch_folder = "E:\\Datasets\\em710\\labels_patches\\"
#################################################################


dsc_loss = DiceLoss(include_background=False, sigmoid=True)


def main():
    image_patch_names = sorted(os.listdir(image_patch_folder))
    label_patch_names = sorted(os.listdir(label_patch_folder))
    image_label_path_pairs = []
    for i in range(len(image_patch_names)):
        image_patch_path = image_patch_folder + image_patch_names[i]
        label_patch_path = label_patch_folder + label_patch_names[i]
        image_label_path_pairs.append([image_patch_path, label_patch_path])

    np.random.shuffle(image_label_path_pairs)
    n_training = int(len(image_label_path_pairs) * 0.8)
    image_label_train_pairs = image_label_path_pairs[:n_training]
    image_label_valid_pairs = image_label_path_pairs[n_training:]


    model = HGM_UNet3D(input_ch=1, output_ch=1, init_feats=16)
    model = model.to(device)

    optimizer = optim.Adam(
        model.parameters(),
        lr=init_learning_rate,
        weight_decay=0.0001,
        eps=1e-5
    )

    scheduler = optim.lr_scheduler.ReduceLROnPlateau(
        optimizer=optimizer,
        mode="min",
        patience=learning_rate_patience,
        factor=learning_rate_factor
    )

    trainSet = Image_Label_train(image_label_pairs=image_label_train_pairs)

    validSet = Image_Label_valid(image_label_pairs=image_label_valid_pairs)

    trainLoader = DataLoader(dataset=trainSet,
                             batch_size=real_batch_size,
                             shuffle=True,
                             drop_last=True,
                             num_workers=4)

    validLoader = DataLoader(dataset=validSet,
                             batch_size=real_batch_size,
                             shuffle=False,
                             drop_last=True,
                             num_workers=4)


    for epoch in range(epochs):
        print("Now is the %d epoch" % (epoch + 1))
        print("The LR now is: ", optimizer.param_groups[0]["lr"])
        model.train()
        #########################################################
        file = Workbook()

        table = file.create_sheet('loss_train_valid_%d' % (epoch + 1), index=0)
        #########################################################
        loss_train = []
        dice_scores = []
        iou_scores = []
        voe_scores = []
        precision_scores = []
        recall_scores = []
        accuracy_scores = []

        optimizer.zero_grad()
        pbar = tqdm(trainLoader, ncols=150)
        for iter, (image, label) in enumerate(pbar):

            image = image.to(device)
            label = label.to(device)
            label_fake = model(image)

            loss = dsc_loss(label_fake, label) / accumulation_steps  # 计算损失

            loss_train.append(loss.item() * accumulation_steps)
            loss.backward()
            if (iter + 1) % accumulation_steps == 0:

                if epoch == 0 and (iter + 1) // accumulation_steps <= 3:
                    print(f"\n{'=' * 60}")
                    print(f"Gradient Check - Epoch {epoch + 1}, Accumulation Step {(iter + 1) // accumulation_steps}")
                    print(f"Loss: {loss.item() * accumulation_steps:.6f}")
                    print("=" * 60)

                    print("=== (first 5 parameters) ===")
                    param_count = 0
                    for name, param in model.named_parameters():
                        if param_count >= 5:
                            remaining = len(list(model.named_parameters())) - 5
                            print(f"... and {remaining} more parameters")
                            break
                        if param.grad is not None:
                            grad = param.grad
                            print(f"{name:40s}: "
                                  f"norm={grad.norm().item():8.4f}, " # norm最值得关注 
                                  f"mean={grad.mean().item():8.4f}, "
                                  f"max={grad.max().item():8.4f}")
                        else:
                            print(f"{name:40s}: grad is None")
                        param_count += 1

                    print("\n======")
                    layer_norms = {}
                    for name, param in model.named_parameters():
                        if param.grad is not None:
                            layer = name.split('.')[0]
                            norm = param.grad.norm().item()
                            if layer not in layer_norms:
                                layer_norms[layer] = []
                            layer_norms[layer].append(norm)

                    for layer, norms in layer_norms.items():
                        print(f"Layer {layer:10s}: "
                              f"mean_norm={np.mean(norms):8.4f}, "
                              f"std_norm={np.std(norms):8.4f}, "
                              f"num_params={len(norms)}")
                    print("=" * 60 + "\n")

                optimizer.step()

                optimizer.zero_grad()

                pbar.set_postfix({
                    'loss': loss.item() * accumulation_steps,
                    'step': f'{(iter + 1) // accumulation_steps}/{(len(trainLoader) + accumulation_steps - 1) // accumulation_steps}'
                })

            pred_prob = torch.sigmoid(label_fake)
            dice_score = calculate_dice(pred_prob, label)
            dice_scores.append(dice_score)
            iou_score = calculate_iou(pred_prob, label)
            iou_scores.append(iou_score)
            voe_score = calculate_voe(pred_prob, label)
            voe_scores.append(voe_score)
            precision_score = calculate_precision(pred_prob, label)
            precision_scores.append(precision_score)
            recall_score = calculate_recall(pred_prob, label)
            recall_scores.append(recall_score)
            accuracy_score = calculate_accuracy(pred_prob, label)
            accuracy_scores.append(accuracy_score)


            pbar.set_description(f"Epoch {epoch + 1} [Batch {iter + 1}/{len(trainLoader)}]")

            table.cell(row=iter + 1, column=1, value=iter + 1)
            table.cell(row=iter + 1, column=2, value=loss.item() * accumulation_steps)

        remaining_steps = len(trainLoader) % accumulation_steps
        if remaining_steps > 0:
            print(f"Processing remaining {remaining_steps} batches...")
            # torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
            optimizer.step()
            optimizer.zero_grad()


        table.cell(row=1, column=3, value=np.mean(loss_train))
        table.cell(row=1, column=5, value=np.mean(dice_scores))
        table.cell(row=1, column=7, value=np.mean(iou_scores))
        table.cell(row=1, column=9, value=np.mean(voe_scores))
        table.cell(row=1, column=11, value=np.mean(precision_scores))
        table.cell(row=1, column=13, value=np.mean(recall_scores))
        table.cell(row=1, column=15, value=np.mean(accuracy_scores))

        ###############################################################
        print(f"Epoch[{epoch + 1}]: Loss_train: {np.mean(loss_train):.4f}, "
              f"Train_Dice: {np.mean(dice_scores):.4f}")


        loss_valid, dice_valid, iou_valid, voe_valid, precision_valid, recall_valid, accuracy_valid = eval(net=model,
                                                                                                           dataloader=validLoader)
        table.cell(row=1, column=4, value=loss_valid)
        table.cell(row=1, column=6, value=dice_valid)
        table.cell(row=1, column=8, value=iou_valid)
        table.cell(row=1, column=10, value=voe_valid)
        table.cell(row=1, column=12, value=np.mean(precision_valid))
        table.cell(row=1, column=14, value=np.mean(recall_valid))
        table.cell(row=1, column=16, value=np.mean(accuracy_valid))
        table.cell(row=1, column=17, value=optimizer.param_groups[0]["lr"])

        print(f"Epoch[{epoch + 1}]: Loss_test: {np.mean(loss_valid):.4f}, "
              f"Test_Dice: {np.mean(dice_valid):.4f}")

        file.save(modelsave_path + "UNet3D_%06d.xlsx" % (epoch + 1))

        if scheduler is not None:
            scheduler.step(loss_valid)
        if epoch % 10 == 0:
            torch.save(model.state_dict(), modelsave_path + "UNet3D_%06d.pth" % (epoch + 1))

        if torch.cuda.is_available():
            allocated = torch.cuda.memory_allocated(device) / 1024 ** 3
            cached = torch.cuda.memory_reserved(device) / 1024 ** 3
            print(f"GPU Memory - Allocated: {allocated:.2f}GB, Cached: {cached:.2f}GB")


if __name__ == '__main__':
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    main()