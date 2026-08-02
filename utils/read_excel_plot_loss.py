import os
import argparse
from pathlib import Path
import openpyxl
from matplotlib import pyplot as plt


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def get_xlsx_filenames_from_folder(input_folder):
    file_list = sorted(os.listdir(input_folder))
    excel_names = []
    for i in file_list:
        if os.path.splitext(i)[1] == ".xlsx":
            excel_names.append(i)
    return excel_names


def parse_args():
    parser = argparse.ArgumentParser(
        description="Plot training and validation loss curves from Excel training records."
    )
    parser.add_argument(
        "--input_dir",
        type=Path,
        default=PROJECT_ROOT / "model-log" / "hgm_unet3d",
        help="Directory containing the Excel training records."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=PROJECT_ROOT / "output" / "training_validation_loss.png",
        help="Path for saving the loss curve."
    )
    return parser.parse_args()


args = parse_args()
excel_folder = args.input_dir.resolve()
output_path = args.output.resolve()
output_path.parent.mkdir(parents=True, exist_ok=True)

excel_names = get_xlsx_filenames_from_folder(excel_folder)

epoch, loss_train, loss_valid = [], [], []
for i in range(len(excel_names)):
    epoch.append(i + 1)
    excel_path = excel_folder / excel_names[i]
    workbook = openpyxl.load_workbook(filename=str(excel_path))
    sheet = workbook["loss_train_valid_%d" % (i + 1)]
    loss_train.append(sheet.cell(row=1, column=3).value)
    loss_valid.append(sheet.cell(row=1, column=4).value)

plt.figure()
plt.plot(epoch, loss_train, "r", linewidth=2, label="Training Loss")
plt.plot(epoch, loss_valid, color="#328DCA", linewidth=2, label="Testing Loss")
# plt.ylim(0, 130)
plt.xlabel("Epoch")
plt.ylabel("Loss")
plt.legend(loc="best", frameon=False)
plt.title("Loss of Training and Testing")
plt.savefig(str(output_path))
plt.show()