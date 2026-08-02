import os
import argparse
from pathlib import Path
import openpyxl
from matplotlib import pyplot as plt


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def get_xlsx_filenames_from_folder(input_folder):
    file_list = sorted(os.listdir(input_folder))
    excel_names = []
    for i in file_list:
        if os.path.splitext(i)[1] == ".xlsx":
            excel_names.append(i)
    return excel_names


def parse_args():
    parser = argparse.ArgumentParser(
        description="Plot training and validation DSC curves from Excel training records."
    )
    parser.add_argument(
        "--input_dir",
        type=Path,
        default=PROJECT_ROOT / "model-log" / "hgm_unet3d",
        help="Directory containing the Excel training records."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=PROJECT_ROOT / "output" / "training_validation_dsc.png",
        help="Path for saving the DSC curve."
    )
    return parser.parse_args()


args = parse_args()
excel_folder = args.input_dir.resolve()
output_path = args.output.resolve()
output_path.parent.mkdir(parents=True, exist_ok=True)

excel_names = get_xlsx_filenames_from_folder(excel_folder)

epoch, dice_train, dice_valid = [], [], []
for i in range(len(excel_names)):
    epoch.append(i + 1)
    excel_path = excel_folder / excel_names[i]
    workbook = openpyxl.load_workbook(filename=str(excel_path))
    sheet = workbook["loss_train_valid_%d" % (i + 1)]
    dice_train.append(sheet.cell(row=1, column=5).value)
    dice_valid.append(sheet.cell(row=1, column=6).value)

plt.figure()
plt.plot(epoch, dice_train, "r", linewidth=2, label="Training DSC")
plt.plot(epoch, dice_valid, color="#328DCA", linewidth=2, label="Testing DSC")
# plt.ylim(0, 130)
plt.xlabel("Epoch")
plt.ylabel("DSC")
plt.legend(loc="best", frameon=False)
plt.title("DSC of Training and Testing")
plt.savefig(str(output_path))
plt.show()